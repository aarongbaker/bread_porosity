"""
Export and Reporting Module for Bread Porosity Analysis Tool
Generates CSV, PDF, and Excel reports with charts and statistics.
"""

import csv
import json
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
import logging

try:
    from .shared_utils import calculate_std_dev
except (ImportError, ValueError):
    from shared_utils import calculate_std_dev

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend

logger = logging.getLogger(__name__)


class ExportEngine:
    """Handle data export in multiple formats: CSV, PDF, Excel."""
    
    def __init__(self, output_dir: str = "./output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_to_csv(self, analyses: List[Dict[str, Any]], 
                     filename: str = "batch_analysis.csv") -> Path:
        """
        Export batch analysis results to CSV.
        
        Args:
            analyses: List of analysis result dictionaries
            filename: Output filename
            
        Returns:
            Path to saved CSV file
        """
        if not analyses:
            logger.warning("No analyses to export")
            return None
        
        output_path = self.output_dir / filename
        
        try:
            # Flatten metrics for CSV
            rows = []
            for analysis in analyses:
                row = {
                    'Image': analysis.get('image_path', ''),
                    'Timestamp': analysis.get('timestamp', datetime.now().isoformat()),
                    'Porosity %': analysis.get('metrics', {}).get('porosity_percent', 0),
                    'Num Holes': analysis.get('metrics', {}).get('num_holes', 0),
                    'Mean Diameter mm': analysis.get('metrics', {}).get('mean_hole_diameter_mm', 0),
                    'Holes per cm²': analysis.get('metrics', {}).get('holes_per_cm2', 0),
                    'Aspect Ratio': analysis.get('metrics', {}).get('mean_aspect_ratio', 0),
                    'Orientation': analysis.get('metrics', {}).get('mean_orientation', 0),
                    'Crumb Brightness CV': analysis.get('metrics', {}).get('crumb_brightness_cv', 0),
                    'Uniformity Grade': analysis.get('metrics', {}).get('uniformity_grade', ''),
                    'Quality Score': analysis.get('metrics', {}).get('quality_score', 0),
                }
                rows.append(row)
            
            # Write CSV
            fieldnames = rows[0].keys()
            with open(output_path, 'w', newline='') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(rows)
            
            logger.info(f"Exported {len(rows)} analyses to CSV: {output_path}")
            return output_path
        
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            raise
    
    def export_to_excel(self, analyses: List[Dict[str, Any]], 
                       filename: str = "batch_analysis.xlsx") -> Path:
        """
        Export batch analysis results to Excel with charts.
        
        Args:
            analyses: List of analysis result dictionaries
            filename: Output filename
            
        Returns:
            Path to saved Excel file
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            return None
        
        if not analyses:
            logger.warning("No analyses to export")
            return None
        
        output_path = self.output_dir / filename
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Analysis Results"
            
            # Headers
            headers = ['Image', 'Timestamp', 'Porosity %', 'Num Holes', 'Mean Diameter mm',
                      'Holes/cm²', 'Aspect Ratio', 'Orientation', 'Crumb Brightness CV',
                      'Uniformity Grade', 'Quality Score']
            ws.append(headers)
            
            # Style header row
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add data
            for analysis in analyses:
                metrics = analysis.get('metrics', {})
                row = [
                    analysis.get('image_path', ''),
                    analysis.get('timestamp', datetime.now().isoformat()),
                    metrics.get('porosity_percent', 0),
                    metrics.get('num_holes', 0),
                    metrics.get('mean_hole_diameter_mm', 0),
                    metrics.get('holes_per_cm2', 0),
                    metrics.get('mean_aspect_ratio', 0),
                    metrics.get('mean_orientation', 0),
                    metrics.get('crumb_brightness_cv', 0),
                    metrics.get('uniformity_grade', ''),
                    metrics.get('quality_score', 0),
                ]
                ws.append(row)
            
            # Auto-width columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add summary sheet
            summary_ws = wb.create_sheet("Summary")
            
            porosities = [a.get('metrics', {}).get('porosity_percent', 0) for a in analyses]
            summary_data = [
                ['Metric', 'Value'],
                ['Total Analyses', len(analyses)],
                ['Mean Porosity %', f"{sum(porosities)/len(porosities):.2f}" if porosities else 0],
                ['Min Porosity %', f"{min(porosities):.2f}" if porosities else 0],
                ['Max Porosity %', f"{max(porosities):.2f}" if porosities else 0],
                ['Std Dev Porosity %', f"{calculate_std_dev(porosities):.2f}" if porosities else 0],
            ]
            
            for row in summary_data:
                summary_ws.append(row)
            
            # Style summary sheet
            for row in summary_ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
            
            wb.save(output_path)
            logger.info(f"Exported {len(analyses)} analyses to Excel: {output_path}")
            return output_path
        
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise
    
    def export_to_pdf(self, analyses: List[Dict[str, Any]], 
                     filename: str = "batch_analysis_report.pdf",
                     title: str = "Bread Porosity Analysis Report",
                     include_charts: bool = True) -> Path:
        """
        Export batch analysis results to PDF report.
        
        Args:
            analyses: List of analysis result dictionaries
            filename: Output filename
            title: Report title
            include_charts: Whether to include chart images
            
        Returns:
            Path to saved PDF file
        """
        if not REPORTLAB_AVAILABLE:
            logger.error("reportlab not installed. Install with: pip install reportlab")
            return None
        
        if not analyses:
            logger.warning("No analyses to export")
            return None
        
        output_path = self.output_dir / filename
        
        try:
            doc = SimpleDocTemplate(str(output_path), pagesize=letter,
                                   rightMargin=0.75*inch, leftMargin=0.75*inch,
                                   topMargin=0.75*inch, bottomMargin=0.75*inch)
            
            story = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#007bff'),
                spaceAfter=30,
                alignment=1  # Center
            )
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 0.2*inch))
            
            # Timestamp
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            story.append(Paragraph(f"<b>Generated:</b> {timestamp}", styles['Normal']))
            story.append(Spacer(1, 0.3*inch))
            
            # Summary statistics
            porosities = [a.get('metrics', {}).get('porosity_percent', 0) for a in analyses]
            summary_data = [
                ['Metric', 'Value'],
                ['Total Analyses', str(len(analyses))],
                ['Mean Porosity %', f"{sum(porosities)/len(porosities):.2f}" if porosities else "N/A"],
                ['Min Porosity %', f"{min(porosities):.2f}" if porosities else "N/A"],
                ['Max Porosity %', f"{max(porosities):.2f}" if porosities else "N/A"],
            ]
            
            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 0.3*inch))
            
            # Detailed results table
            story.append(Paragraph("Detailed Results", styles['Heading2']))
            story.append(Spacer(1, 0.15*inch))
            
            # Create detailed table
            table_data = [
                ['Image', 'Porosity %', 'Holes', 'Diameter mm', 'Quality']
            ]
            
            for analysis in analyses[:20]:  # Limit to first 20 for readability
                metrics = analysis.get('metrics', {})
                image_name = Path(analysis.get('image_path', '')).name
                table_data.append([
                    image_name[:30],
                    f"{metrics.get('porosity_percent', 0):.1f}",
                    str(metrics.get('num_holes', 0)),
                    f"{metrics.get('mean_hole_diameter_mm', 0):.2f}",
                    metrics.get('uniformity_grade', '-'),
                ])
            
            if len(analyses) > 20:
                table_data.append(['... and ' + str(len(analyses) - 20) + ' more', '', '', '', ''])
            
            details_table = Table(table_data, colWidths=[2*inch, 1.2*inch, 1*inch, 1.2*inch, 1*inch])
            details_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28a745')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            story.append(details_table)
            
            # Build PDF
            doc.build(story)
            logger.info(f"Exported PDF report: {output_path}")
            return output_path
        
        except Exception as e:
            logger.error(f"Error exporting to PDF: {e}")
            raise
    
    def create_summary_charts(self, analyses: List[Dict[str, Any]]) -> Dict[str, Path]:
        """
        Create summary charts and return paths.
        
        Args:
            analyses: List of analysis result dictionaries
            
        Returns:
            Dictionary mapping chart names to file paths
        """
        if not analyses:
            logger.warning("No analyses for charts")
            return {}
        
        chart_paths = {}
        porosities = [a.get('metrics', {}).get('porosity_percent', 0) for a in analyses]
        hole_counts = [a.get('metrics', {}).get('num_holes', 0) for a in analyses]
        
        try:
            # Porosity trend chart
            fig, ax = plt.subplots(figsize=(10, 6))
            ax.plot(porosities, marker='o', linewidth=2, markersize=6, color='#007bff')
            ax.axhline(y=sum(porosities)/len(porosities), color='r', linestyle='--', label='Mean')
            ax.set_xlabel('Analysis Number')
            ax.set_ylabel('Porosity %')
            ax.set_title('Porosity Trend')
            ax.grid(True, alpha=0.3)
            ax.legend()
            
            path = self.output_dir / "chart_porosity_trend.png"
            plt.savefig(path, dpi=150, bbox_inches='tight')
            plt.close()
            chart_paths['porosity_trend'] = path
            
            # Hole count chart
            fig, ax = plt.subplots(figsize=(10, 6))
            ax.bar(range(len(hole_counts)), hole_counts, color='#28a745', alpha=0.7)
            ax.set_xlabel('Analysis Number')
            ax.set_ylabel('Hole Count')
            ax.set_title('Hole Count Distribution')
            ax.grid(True, alpha=0.3, axis='y')
            
            path = self.output_dir / "chart_hole_count.png"
            plt.savefig(path, dpi=150, bbox_inches='tight')
            plt.close()
            chart_paths['hole_count'] = path
            
            logger.info(f"Created {len(chart_paths)} summary charts")
            return chart_paths
        
        except Exception as e:
            logger.error(f"Error creating charts: {e}")
            return {}
    

